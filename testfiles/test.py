from xlrd import open_workbook

wb = open_workbook('test_xlsx.xlsx')
sheet = wb.sheet_by_name("Sheet1")
# cell = sheet.cell(1, 1)
# print(cell.xf_index)
# fmt = wb.xf_list[cell.xf_index]
# fmt.dump()

from openpyxl import load_workbook
wb = load_workbook('test_xlsx.xlsx', data_only=True)
sh = wb['Sheet1']
print(sh['B2'].value)
print(sh['B2'].fill.start_color)